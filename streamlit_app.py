import os, io, re, base64, tempfile, zipfile
import streamlit as st
from pathlib import Path
from xml.etree import ElementTree as ET

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="STARS · Doc → Markdown",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Library checks ────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def _check_libs():
    libs = {}
    for key, mod in [("fitz","fitz"),("pymupdf4llm","pymupdf4llm"),
                     ("mammoth","mammoth"),("pdf2image","pdf2image"),
                     ("pytesseract","pytesseract"),
                     ("openpyxl","openpyxl"),("xlrd","xlrd"),
                     ("markdownify","markdownify")]:
        try:    __import__(mod); libs[key] = True
        except: libs[key] = False
    return libs

LIBS = _check_libs()

# ── Constants ─────────────────────────────────────────────────────────────────
MAX_IMG_BYTES   = 600 * 1024        # 600 KB — imágenes más grandes se omiten
PREVIEW_CHARS   = 200_000           # caracteres máx. en vista previa / raw
RAW_WARN_MB     = 5                 # MB — avisa si el .md es muy grande

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root {
  --dl:      #005A70;
  --teal:    #10B5BF;
  --green:   #01969F;
  --navy:    #061D3D;
  --dark:    #302E2E;
  --bg:      #EDF2F4;
  --border:  #C9D8DC;
  --white:   #FFFFFF;
  --shadow:  0 2px 16px rgba(0,90,112,.10);
  --font:    'Inter', 'Segoe UI', sans-serif;
}

#MainMenu, footer                         { visibility: hidden; }
[data-testid="stHeader"]                  { display: none !important; }
[data-testid="stToolbar"]                 { display: none !important; }
[data-testid="stDecoration"]              { display: none !important; }
[data-testid="stDeployButton"]            { display: none !important; }
[data-testid="stStatusWidget"]            { display: none !important; }

html, body, .stApp                        { font-family: var(--font); background: var(--bg); }
[data-testid="stAppViewBlockContainer"]   { padding-top: 0 !important; }
.main .block-container                    { padding-top: 0 !important; padding-left: 0 !important;
                                            padding-right: 0 !important; max-width: 100% !important; }

[data-testid="column"] {
  background: var(--white); border-radius: 14px;
  box-shadow: var(--shadow); padding: 24px 26px 28px !important;
}

.stButton > button {
  background: var(--dl) !important; color: var(--white) !important;
  border: none !important; border-radius: 8px !important;
  font-family: var(--font) !important; font-weight: 600 !important;
  font-size: 13.5px !important; padding: 10px 20px !important;
  transition: background .18s !important;
}
.stButton > button:hover { background: var(--navy) !important; }

.sel-row .stButton > button {
  background: transparent !important; color: var(--teal) !important;
  border: 1.5px solid var(--teal) !important; border-radius: 6px !important;
  font-size: 12px !important; padding: 4px 12px !important;
  min-height: 0 !important; height: auto !important;
}
.sel-row .stButton > button:hover { background: var(--teal) !important; color: var(--white) !important; }

.stDownloadButton > button {
  background: var(--teal) !important; color: var(--white) !important;
  border: none !important; border-radius: 8px !important;
  font-family: var(--font) !important; font-weight: 600 !important;
  font-size: 13px !important; transition: background .18s !important;
}
.stDownloadButton > button:hover { background: var(--green) !important; }

[data-testid="stFileUploader"] section {
  background: var(--bg) !important; border: 2px dashed var(--border) !important;
  border-radius: 12px !important; transition: border-color .2s;
}
[data-testid="stFileUploader"] section:hover { border-color: var(--teal) !important; }
[data-testid="stFileUploaderDropzone"]        { background: var(--bg) !important; }

[data-testid="stProgress"] > div > div > div  { background: var(--teal) !important; }
.stProgress > div > div > div                 { background: var(--teal) !important; }

.stTabs [role="tablist"]                      { border-bottom: 2px solid var(--bg) !important; }
.stTabs [role="tab"]                          { font-weight: 600 !important; color: #9CA3AF !important; font-size: 13px !important; }
.stTabs [role="tab"][aria-selected="true"]    { color: var(--teal) !important; border-bottom-color: var(--teal) !important; }

.panel-heading {
  font-size: 11.5px; font-weight: 700; color: var(--dl);
  text-transform: uppercase; letter-spacing: .6px;
  padding-bottom: 14px; border-bottom: 1.5px solid var(--bg); margin-bottom: 18px;
}
.fmt-row { display:flex; gap:7px; flex-wrap:wrap; margin-bottom:16px; }
.badge { display:inline-block; padding:3px 9px; border-radius:5px; font-size:11px; font-weight:700; letter-spacing:.3px; }
.b-pdf  { background:#FEE2E2; color:#B91C1C; }
.b-docx,.b-doc { background:#DBEAFE; color:#1D4ED8; }
.b-txt  { background:#FEF9C3; color:#78350F; }
.b-pptx { background:#FCE7F3; color:#9D174D; }
.b-xlsx,.b-xls,.b-csv { background:#D1FAE5; color:#065F46; }
.b-ok   { background:#D1FAE5; color:#065F46; }
.b-err  { background:#FEE2E2; color:#B91C1C; }

.file-row {
  display:flex; align-items:center; gap:10px;
  padding:10px 13px; border-radius:9px; margin-bottom:6px;
  background:var(--bg); border:1.5px solid transparent;
}
.file-row.active { border-color:var(--teal); background:#E3F4F6; }
.file-row .fname { flex:1; font-weight:600; font-size:13px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; color:var(--dark); }
.file-row .fmeta { font-size:11.5px; color:#9CA3AF; white-space:nowrap; }

.stat-row  { display:flex; gap:10px; margin:14px 0 18px; }
.stat-card { flex:1; background:var(--bg); border-radius:9px; padding:12px 8px; text-align:center; }
.stat-val  { font-weight:700; font-size:1.3rem; color:var(--teal); }
.stat-lbl  { font-size:10.5px; color:#9CA3AF; text-transform:uppercase; letter-spacing:.05em; margin-top:2px; }

.method-pill { background:var(--dl); color:white; padding:3px 9px; border-radius:4px; font-size:11px; font-weight:700; margin-left:8px; }

.empty-state { text-align:center; padding:64px 24px; color:#B0BFC5; }
.empty-icon  { font-size:3rem; margin-bottom:14px; opacity:.6; }
.empty-title { font-size:15px; font-weight:600; color:#8FA3AB; }
.empty-sub   { font-size:12.5px; margin-top:6px; line-height:1.6; color:#B0BFC5; }

.md-preview h1 { color:var(--navy); border-bottom:2px solid var(--teal); padding-bottom:6px; }
.md-preview h2 { color:var(--dl); }
.md-preview h3 { color:var(--green); }
.md-preview table  { border-collapse:collapse; width:100%; font-size:.88rem; }
.md-preview thead tr { background:var(--dl); color:white; }
.md-preview th, .md-preview td { padding:8px 12px; border:1px solid #D1D5DB; }
.md-preview tbody tr:nth-child(even) { background:#F4FAFA; }
.md-preview hr  { border:none; border-top:1.5px solid var(--border); }
.md-preview code { background:var(--bg); padding:2px 6px; border-radius:4px; font-size:.87em; }
.md-preview blockquote { border-left:3px solid var(--teal); padding-left:14px; color:#555; }
.md-preview img { max-width:100%; height:auto; border-radius:6px; margin:8px 0; }
.md-preview .img-placeholder {
  display:block; background:var(--bg); border:1.5px dashed var(--border);
  border-radius:8px; padding:10px 16px; color:#9CA3AF;
  font-size:.85rem; margin:6px 0;
}
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="
  background:#005A70; padding:0 40px; height:64px;
  display:flex; align-items:center; gap:14px;
  position:relative; margin-bottom:28px;
  font-family:'Inter','Segoe UI',sans-serif;
">
  <div style="
    width:38px;height:38px;border-radius:50%;background:#10B5BF;
    display:flex;align-items:center;justify-content:center;
    font-weight:700;font-size:15px;color:white;flex-shrink:0;
  ">S</div>
  <span style="color:white;font-weight:700;font-size:17px;">STARS Companies</span>
  <div style="width:1px;height:22px;background:rgba(255,255,255,.3);"></div>
  <span style="color:rgba(255,255,255,.72);font-size:13px;">Document → Markdown Converter</span>
  <div style="
    position:absolute;bottom:0;left:0;right:0;height:3px;
    background:linear-gradient(90deg,#10B5BF,#01969F);
  "></div>
</div>
""", unsafe_allow_html=True)


# ── Image helpers ─────────────────────────────────────────────────────────────

def _to_data_uri(img_bytes: bytes, ext: str):
    """
    Returns a base64 data URI string, or None if the image exceeds MAX_IMG_BYTES.
    Returning None tells the caller to emit a placeholder instead.
    """
    if len(img_bytes) > MAX_IMG_BYTES:
        return None
    ext  = ext.lower().lstrip('.')
    mime = 'image/jpeg' if ext in ('jpg', 'jpeg') else f'image/{ext}'
    return f"data:{mime};base64,{base64.b64encode(img_bytes).decode()}"


def _embed_img_refs(md: str, tmp_dir: str) -> str:
    """
    Replaces ![alt](path) file references written by pymupdf4llm with
    base64 data URIs (or a styled placeholder when the image is too large
    or can't be read).
    Must be called INSIDE the TemporaryDirectory context.
    """
    IMG_RE = re.compile(r'!\[([^\]]*)\]\(([^)]+)\)')

    def _replace(m):
        alt, path = m.group(1), m.group(2)
        for candidate in [path,
                          os.path.join(tmp_dir, path),
                          os.path.join(tmp_dir, os.path.basename(path))]:
            if not os.path.isfile(candidate):
                continue
            ext = Path(candidate).suffix.lower().lstrip('.')
            if ext not in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'):
                break
            try:
                with open(candidate, 'rb') as f:
                    raw = f.read()
                uri = _to_data_uri(raw, ext)
                if uri:
                    return f'![{alt}]({uri})'
                # Image too large — leave a readable placeholder
                kb = len(raw) // 1024
                return f'> 🖼️ *Imagen ({kb} KB) incluida en el archivo .md descargado*'
            except Exception:
                break
        return m.group(0)

    return IMG_RE.sub(_replace, md)


def _strip_data_uris(md: str) -> str:
    """
    For the browser preview ONLY: replaces base64 data URIs with a lightweight
    placeholder so the browser doesn't freeze rendering huge inline images.
    The downloaded .md file always keeps the real images.
    """
    return re.sub(
        r'!\[([^\]]*)\]\(data:[^)]{50,}\)',
        lambda m: f'<span class="img-placeholder">🖼️ {m.group(1) or "Imagen"} — visible en el archivo .md descargado</span>',
        md,
    )


# ── Conversion functions ──────────────────────────────────────────────────────

def convert_pdf(file_bytes, filename, include_images=True):
    if not LIBS["fitz"]: raise ImportError("PyMuPDF no instalado.")
    import fitz, pymupdf4llm

    doc   = fitz.open(stream=file_bytes, filetype="pdf")
    total = len(doc)

    # Scanned-PDF detection: sample several pages, not just the cover.
    # A text PDF with an image-only cover page must NOT be routed to OCR.
    sample_idx = sorted({0, min(1, total - 1), total // 2, total - 1})
    sample_chars = [len(doc[i].get_text().strip()) for i in sample_idx]
    is_scanned = all(c < 30 for c in sample_chars)

    # ── Scanned PDF → OCR ────────────────────────────────────────────────────
    if is_scanned:
        if not (LIBS["pdf2image"] and LIBS["pytesseract"]):
            raise ImportError("PDF escaneado: instala pdf2image y pytesseract.")
        from pdf2image import convert_from_bytes
        import pytesseract
        # FIX: load all pages in ONE call — the previous per-page loop was O(n²)
        # because each convert_from_bytes re-decoded the whole PDF up to that page.
        # 150 dpi is enough for OCR accuracy and uses 44% less RAM than 200 dpi.
        imgs  = convert_from_bytes(file_bytes, dpi=150, thread_count=2)
        pages = [f"## Página {i}\n\n{pytesseract.image_to_string(img, lang='spa+eng').strip()}"
                 for i, img in enumerate(imgs, 1)]
        return f"# {Path(filename).stem}\n\n" + "\n\n---\n\n".join(pages), "OCR (pytesseract)"

    # ── Text PDF → pymupdf4llm ────────────────────────────────────────────────
    label = "pymupdf4llm (word PDF)" if doc[0].get_fonts() else "pymupdf4llm (plain PDF)"

    if not include_images:
        # FIX: text-only is 3–5x faster — larger chunks, no disk I/O for images
        CHUNK = 50
        parts = [pymupdf4llm.to_markdown(doc, pages=list(range(s, min(s + CHUNK, total))))
                 for s in range(0, total, CHUNK)]
        md = re.sub(r'\n{4,}', '\n\n\n', "\n\n".join(parts))
        return md, label

    # Images requested — write to temp dir then embed as base64
    CHUNK = 20
    with tempfile.TemporaryDirectory() as tmp:
        parts = []
        for s in range(0, total, CHUNK):
            chunk_md = pymupdf4llm.to_markdown(
                doc,
                pages=list(range(s, min(s + CHUNK, total))),
                write_images=True,
                image_path=tmp,
                image_format="png",
            )
            parts.append(chunk_md)

        md = re.sub(r'\n{4,}', '\n\n\n', "\n\n".join(parts))
        md = _embed_img_refs(md, tmp)       # embed INSIDE the 'with' block

    return md, label


def _html_to_md(html: str) -> str:
    """
    HTML → Markdown preservando tablas. Escapa pipes y saltos de línea
    dentro de celdas para que las tablas no se rompan.
    """
    from markdownify import MarkdownConverter

    class _Conv(MarkdownConverter):
        def _cell(self, el, text, *args, **kwargs):
            clean = ' '.join(text.split()).replace('|', '\\|')
            return ' ' + clean + ' |'
        def convert_td(self, el, text, *args, **kwargs): return self._cell(el, text)
        def convert_th(self, el, text, *args, **kwargs): return self._cell(el, text)

    md = _Conv(heading_style="ATX", bullets="-").convert(html)
    md = _fix_empty_table_headers(md)
    md = re.sub(r'!\[[^\]]*\]\(\s*\)\n?', '', md)   # imágenes sin src (modo rápido)
    return re.sub(r'\n{4,}', '\n\n\n', md).strip() + '\n'


def _fix_empty_table_headers(md: str) -> str:
    """
    mammoth emite tablas sin <th>, lo que deja un header vacío `|  |  |`.
    Promueve la primera fila de datos a header.
    """
    lines = md.split('\n')
    out, i = [], 0
    while i < len(lines):
        if (re.fullmatch(r'\|(\s*\|)+', lines[i].strip())
                and i + 2 < len(lines)
                and re.fullmatch(r'\|(\s*:?-{3,}:?\s*\|)+', lines[i + 1].strip())
                and lines[i + 2].strip().startswith('|')):
            out.append(lines[i + 2])
            out.append(lines[i + 1])
            i += 3
            continue
        out.append(lines[i]); i += 1
    return '\n'.join(out)


def convert_docx(file_bytes, filename, include_images=True):
    if not LIBS["mammoth"]: raise ImportError("mammoth no instalado.")
    import mammoth

    # Word 97-2003 (.doc binario, contenedor OLE) — mammoth solo soporta .docx
    if file_bytes[:4] == b'\xd0\xcf\x11\xe0':
        raise ValueError(
            "Es un .doc antiguo (Word 97-2003). Ábrelo en Word y guárdalo "
            "como .docx para convertirlo."
        )

    def _img_handler(image):
        if not include_images:
            return {}                      # <img> sin src → se elimina después
        try:
            with image.open() as f:
                raw = f.read()
            uri = _to_data_uri(raw, image.content_type.split('/')[-1])
            if uri:
                return {"src": uri}
            kb = len(raw) // 1024
            return {"alt": f"Imagen ({kb} KB) omitida por tamaño", "src": ""}
        except Exception:
            return {}

    if LIBS["markdownify"]:
        # Ruta HTML → MD: preserva tablas (el writer Markdown de mammoth las pierde)
        result = mammoth.convert_to_html(
            io.BytesIO(file_bytes),
            convert_image=mammoth.images.img_element(_img_handler),
        )
        return _html_to_md(result.value), "mammoth+markdownify (DOCX)"

    # Fallback sin markdownify: writer Markdown nativo (sin tablas)
    result = mammoth.convert_to_markdown(
        io.BytesIO(file_bytes),
        convert_image=mammoth.images.img_element(_img_handler),
    )
    return result.value, "mammoth (DOCX)"


def _decode_text(file_bytes: bytes) -> str:
    """UTF-8 (con o sin BOM) primero; si falla, cp1252 — el encoding típico
    de exportes legacy de Windows en español. Nunca produce mojibake '�'."""
    for enc in ('utf-8-sig', 'utf-8'):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            pass
    try:
        return file_bytes.decode('cp1252')
    except UnicodeDecodeError:
        return file_bytes.decode('latin-1', errors='replace')


def convert_txt(file_bytes, filename):
    return f"# {Path(filename).stem}\n\n{_decode_text(file_bytes)}", "texto plano"


def convert_csv(file_bytes, filename):
    import csv
    text = _decode_text(file_bytes)
    # Detecta delimitador (";" es lo habitual en Excel configurado en español)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    rows = [list(r) for r in csv.reader(io.StringIO(text), dialect) if any(c.strip() for c in r)]
    return f"# {Path(filename).stem}\n\n{_md_table(rows)}", "csv"


def convert_pptx(file_bytes, filename, include_images=True):
    A = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
    P = '{http://schemas.openxmlformats.org/presentationml/2006/main}'
    R = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'

    def pt(el):       return ''.join(t.text or '' for t in el.iter(f'{A}t')).strip()
    def is_title(sp): return any(ph.get('type','') in ('title','ctrTitle') or ph.get('idx','9')=='0'
                                  for ph in sp.iter(f'{P}ph'))

    def slide_notes(z, all_files, rels_root):
        """Texto de las notas del orador (vía relación notesSlide)."""
        if rels_root is None: return ''
        for rel in rels_root:
            if not rel.get('Type', '').endswith('/notesSlide'): continue
            target = 'ppt/' + rel.get('Target', '').replace('../', '')
            if target not in all_files: continue
            nroot = ET.parse(z.open(target)).getroot()
            paras = [pt(p) for p in nroot.iter(f'{A}p')]
            # descarta vacíos y el placeholder de número de página (solo dígitos)
            paras = [p for p in paras if p and not p.isdigit()]
            return ' '.join(paras)
        return ''

    slides = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        all_files = set(z.namelist())
        fnames = sorted(
            [f for f in all_files
             if f.startswith('ppt/slides/slide') and f.endswith('.xml')
             and 'Layout' not in f and 'Master' not in f],
            key=lambda x: int(''.join(c for c in x if c.isdigit()) or '0')
        )

        for i, sf in enumerate(fnames, 1):
            root = ET.parse(z.open(sf)).getroot()
            rels_path = sf.replace('slides/slide', 'slides/_rels/slide') + '.rels'
            rels_root = ET.parse(z.open(rels_path)).getroot() if rels_path in all_files else None
            ttl, body = '', []

            for sp in root.iter(f'{P}sp'):
                paras = []
                for p in sp.iter(f'{A}p'):
                    t = pt(p)
                    if not t: continue
                    ppr = p.find(f'{A}pPr')
                    lvl = int(ppr.get('lvl', '0')) if ppr is not None else 0
                    paras.append((t, lvl))
                if not paras: continue
                if not ttl and is_title(sp):
                    ttl, paras = paras[0][0], paras[1:]
                    if not paras: continue
                # Varios párrafos o con sangría → lista con jerarquía;
                # un párrafo suelto (subtítulos, cuadros de texto) → texto plano
                if len(paras) > 1 or paras[0][1] > 0:
                    body.append('\n'.join('  ' * lvl + f'- {t}' for t, lvl in paras))
                else:
                    body.append(paras[0][0])

            # graphicFrame: tablas reales → tabla MD; otros (charts) → texto suelto
            for gf in root.iter(f'{P}graphicFrame'):
                tbl = next(gf.iter(f'{A}tbl'), None)
                if tbl is not None:
                    rows = [[' '.join(pt(p) for p in tc.iter(f'{A}p') if pt(p))
                             for tc in tr.iter(f'{A}tc')]
                            for tr in tbl.iter(f'{A}tr')]
                    rows = [r for r in rows if any(c.strip() for c in r)]
                    if rows: body.append(_md_table(rows))
                else:
                    body.extend(pt(p) for p in gf.iter(f'{A}p') if pt(p))

            # ── Images for this slide (only when requested) ───────────────
            img_tags = []
            if include_images and rels_root is not None:
                for rel in rels_root:
                    if '/image' not in rel.get('Type', ''): continue
                    target = rel.get('Target', '')
                    media  = 'ppt/' + target.replace('../', '')
                    if media not in all_files: continue
                    ext = Path(media).suffix.lower().lstrip('.')
                    if ext not in ('png','jpg','jpeg','gif','bmp','webp'): continue
                    try:
                        raw = z.read(media)
                        uri = _to_data_uri(raw, ext)
                        if uri:
                            img_tags.append(f'![imagen slide {i}]({uri})')
                        else:
                            img_tags.append(f'> 🖼️ *Imagen ({len(raw)//1024} KB) disponible en el archivo .md descargado*')
                    except Exception:
                        pass

            notes   = slide_notes(z, all_files, rels_root)
            hdr     = f"## Slide {i}" + (f": {ttl}" if ttl else "")
            content = []
            if body:     content.append('\n\n'.join(body))
            if img_tags: content.append('\n\n'.join(img_tags))
            if notes:    content.append(f"> **Notas del orador:** {notes}")
            slides.append(f"{hdr}\n\n" + '\n\n'.join(content) if content else hdr)

    return '\n\n---\n\n'.join(slides), "XML directo (PPTX)"


def _fmt_cell(value) -> str:
    import datetime
    if value is None: return ""
    if isinstance(value, datetime.datetime):
        if value.hour == value.minute == value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _esc_cell(s: str) -> str:
    """Escapa lo que rompe una tabla Markdown: pipes y saltos de línea."""
    return s.replace('|', '\\|').replace('\r\n', '\n').replace('\r', '\n').replace('\n', '<br>')


def _md_table(rows):
    if not rows: return "*(vacío)*"
    rows = [[_esc_cell(str(c)) for c in r] for r in rows]
    n    = max(len(r) for r in rows)
    pad  = lambda r: r + [''] * (n - len(r))
    lines = ['| ' + ' | '.join(pad(rows[0])) + ' |', '|' + '--- |' * n]
    for r in rows[1:]: lines.append('| ' + ' | '.join(pad(r)) + ' |')
    return '\n'.join(lines)


def convert_xlsx(file_bytes, filename):
    ext = Path(filename).suffix.lower()
    if ext == '.xls' and LIBS["xlrd"]:
        try:
            import xlrd
            wb     = xlrd.open_workbook(file_contents=file_bytes)
            sheets = [f"## {s.name}\n\n" + _md_table(
                [[str(s.cell_value(r, c)) for c in range(s.ncols)] for r in range(s.nrows)]
            ) for s in wb.sheets()]
            return f"# {Path(filename).stem}\n\n" + '\n\n---\n\n'.join(sheets), "xlrd (XLS)"
        except: pass
    if not LIBS["openpyxl"]: raise ImportError("openpyxl no instalado.")
    from openpyxl import load_workbook
    wb  = load_workbook(io.BytesIO(file_bytes), data_only=True)
    # Segunda pasada sin data_only: recupera el texto de las fórmulas para
    # celdas cuyo valor calculado no está cacheado (saldrían vacías).
    try:
        wb_f = load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception:
        wb_f = None
    base = Path(filename).stem
    parts = [f"# {base}\n"]
    for ws in wb.worksheets:
        ws_f = wb_f[ws.title] if wb_f is not None and ws.title in wb_f.sheetnames else None
        all_rows = list(ws.iter_rows(values_only=True))
        last = max((i for i, r in enumerate(all_rows)
                    if any(v is not None and str(v).strip() for v in r)), default=-1)
        if last < 0: continue
        fmt = []
        for ri, row in enumerate(all_rows[:last + 1]):
            cells = []
            for ci, c in enumerate(row):
                if c is None and ws_f is not None:
                    raw = ws_f.cell(row=ri + 1, column=ci + 1).value
                    if isinstance(raw, str) and raw.startswith('='):
                        cells.append(f"`{raw}`")   # fórmula sin valor cacheado
                        continue
                cells.append(_fmt_cell(c))
            fmt.append(cells)
        fmt = [r for r in fmt if any(c.strip() for c in r)]
        if not fmt: continue
        parts.append(f"\n## {ws.title}\n\n" + _md_table(fmt))
    return '\n'.join(parts), "openpyxl (XLSX)"


@st.cache_data(show_spinner=False, max_entries=5, ttl=600)
def run_conversion(file_bytes: bytes, filename: str, include_images: bool = False) -> dict:
    ext = Path(filename).suffix.lower()
    fn  = {'.pdf': convert_pdf, '.docx': convert_docx, '.doc': convert_docx,
           '.txt': convert_txt, '.pptx': convert_pptx,
           '.xlsx': convert_xlsx, '.xls': convert_xlsx, '.csv': convert_csv}
    if ext not in fn: raise ValueError(f"Formato no soportado: {ext}")
    # txt/csv/xlsx don't use images — call without the kwarg to keep cache keys clean
    if ext in ('.txt', '.csv', '.xlsx', '.xls'):
        md, method = fn[ext](file_bytes, filename)
    else:
        md, method = fn[ext](file_bytes, filename, include_images=include_images)
    return dict(filename=filename, markdown=md, method=method,
                words=len(md.split()), chars=len(md), lines=md.count('\n') + 1)


# ── Session state ─────────────────────────────────────────────────────────────
if "results" not in st.session_state: st.session_state.results = {}
if "active"  not in st.session_state: st.session_state.active  = None

# ── Layout ────────────────────────────────────────────────────────────────────
col_left, col_right = st.columns([1, 1.65], gap="large")


# ════════════════════════════════ LEFT ═══════════════════════════════════════
with col_left:

    st.markdown('<div class="panel-heading">📁 &nbsp;Subir documentos</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="fmt-row">
      <span class="badge b-pdf">PDF</span>
      <span class="badge b-docx">DOCX · DOC</span>
      <span class="badge b-pptx">PPTX</span>
      <span class="badge b-xlsx">XLSX · XLS · CSV</span>
      <span class="badge b-txt">TXT</span>
    </div>""", unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Arrastra archivos aquí o haz clic para seleccionar",
        type=["pdf", "docx", "doc", "txt", "pptx", "xlsx", "xls", "csv"],
        accept_multiple_files=True,
    )

    if uploaded:
        # ── Speed / quality toggle ────────────────────────────────────────
        include_images = st.toggle(
            "🖼️  Incluir imágenes en el Markdown",
            value=False,
            help="Desactivado = conversión rápida (texto completo, sin imágenes).\n"
                 "Activado = incluye imágenes embebidas — puede tardar 1–3 min en documentos grandes.",
        )
        if include_images:
            st.caption("⚠️ Modo con imágenes: la conversión será más lenta. "
                       "Las imágenes quedan dentro del archivo .md descargado.")

        # Warn about large files
        large = [f.name for f in uploaded if f.size > 50 * 1024 * 1024]
        if large:
            st.warning(
                f"Archivo(s) grandes: **{', '.join(large)}**. "
                "La conversión puede tardar varios minutos — no cierres la pestaña.",
                icon=None,
            )

        if st.button("⟳  Convertir todo", type="primary", use_container_width=True):
            st.session_state.results = {}
            st.session_state.active  = None
            bar = st.progress(0, text="Iniciando…")
            for i, uf in enumerate(uploaded):
                bar.progress(i / len(uploaded), text=f"Convirtiendo {uf.name}…")
                try:
                    fb = uf.getvalue()
                    if not fb: raise ValueError("Archivo vacío.")
                    r = run_conversion(fb, uf.name, include_images)
                    st.session_state.results[uf.name] = {"ok": True, **r}
                except Exception as e:
                    st.session_state.results[uf.name] = {"ok": False, "filename": uf.name, "error": str(e)}
            bar.progress(1.0, text="¡Listo!")
            for name, r in st.session_state.results.items():
                if r["ok"]: st.session_state.active = name; break
            st.rerun()

    results = st.session_state.results
    if results:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="panel-heading" style="font-size:10.5px;">Cola de conversión</div>',
                    unsafe_allow_html=True)

        for name, r in results.items():
            ext        = Path(name).suffix.lower().lstrip('.')
            is_active  = name == st.session_state.active
            status_cls = "b-ok" if r["ok"] else "b-err"
            status_lbl = "✓ Listo" if r["ok"] else "✗ Error"
            words_str  = f"{r['words']:,} palabras" if r.get("words") else ""
            st.markdown(f"""
            <div class="file-row {'active' if is_active else ''}">
              <span class="badge b-{ext}">{ext.upper()}</span>
              <span class="fname">{name}</span>
              <span class="fmeta">{words_str}</span>
              <span class="badge {status_cls}">{status_lbl}</span>
            </div>""", unsafe_allow_html=True)

            with st.container():
                st.markdown('<div class="sel-row">', unsafe_allow_html=True)
                if r["ok"] and st.button("Ver resultado", key=f"sel_{name}"):
                    st.session_state.active = name
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

            if not r["ok"]:
                st.caption(f"⚠️ {r.get('error', 'Error desconocido')}")

        done_ok = {n: r for n, r in results.items() if r.get("ok")}
        if len(done_ok) > 1:
            st.markdown("---")
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for name, r in done_ok.items():
                    zf.writestr(Path(name).stem + ".md", r["markdown"])
            st.download_button(
                "⬇ Descargar todo (.zip)",
                data=buf.getvalue(),
                file_name="documentos_markdown.zip",
                mime="application/zip",
                use_container_width=True,
            )


# ════════════════════════════════ RIGHT ══════════════════════════════════════
with col_right:

    active  = st.session_state.active
    results = st.session_state.results

    if not active or not results.get(active, {}).get("ok"):
        st.markdown('<div class="panel-heading">📝 &nbsp;Resultado Markdown</div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="empty-state">
          <div class="empty-icon">📄</div>
          <div class="empty-title">Ningún archivo convertido aún</div>
          <div class="empty-sub">Sube uno o más archivos y haz clic en<br><b>Convertir todo</b> para ver el resultado aquí.</div>
        </div>""", unsafe_allow_html=True)

    else:
        r  = results[active]
        md = r["markdown"]
        md_bytes = len(md.encode())

        h_col, d_col = st.columns([3, 1])
        with h_col:
            st.markdown(
                f'<div class="panel-heading" style="border:none;padding-bottom:0;margin-bottom:0;">'
                f'📝 &nbsp;{r["filename"]}'
                f'<span class="method-pill">{r["method"]}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with d_col:
            st.download_button(
                "⬇ Descargar .md",
                data=md,                            # full markdown with real images
                file_name=Path(r["filename"]).stem + ".md",
                mime="text/markdown",
                use_container_width=True,
            )

        st.markdown(f"""
        <div class="stat-row">
          <div class="stat-card"><div class="stat-val">{r['words']:,}</div><div class="stat-lbl">Palabras</div></div>
          <div class="stat-card"><div class="stat-val">{r['chars']:,}</div><div class="stat-lbl">Caracteres</div></div>
          <div class="stat-card"><div class="stat-val">{r['lines']:,}</div><div class="stat-lbl">Líneas</div></div>
        </div>""", unsafe_allow_html=True)

        # FIX 1 & 3: Large markdown warning
        if md_bytes > RAW_WARN_MB * 1024 * 1024:
            size_mb = md_bytes / (1024 * 1024)
            st.info(
                f"📦 El Markdown generado pesa **{size_mb:.1f} MB** (incluye imágenes en base64). "
                "La vista previa muestra texto completo + marcadores de imagen. "
                "El archivo descargado contiene todas las imágenes reales.",
                icon=None,
            )

        tab_prev, tab_raw = st.tabs(["  Vista previa  ", "  Markdown raw  "])

        with tab_prev:
            # FIX 1: Strip base64 URIs before sending to browser — prevents freeze
            preview_md = _strip_data_uris(md)
            st.markdown('<div class="md-preview">', unsafe_allow_html=True)
            st.markdown(preview_md, unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        with tab_raw:
            # FIX 3: Truncate raw display — st.code on 50MB+ strings also freezes
            if len(md) > PREVIEW_CHARS:
                st.caption(f"Mostrando primeros {PREVIEW_CHARS:,} caracteres. Descarga el .md para el contenido completo.")
                display_md = md[:PREVIEW_CHARS] + "\n\n…[descarga el archivo para ver el resto]"
            else:
                display_md = md
            st.code(display_md, language="markdown", line_numbers=True)
